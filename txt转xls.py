
# -*- coding: UTF-8 -*-
import os
import time
 
import openpyxl as openpyxl
Parser_version = 'version_1.0'
print Parser_version
items = os.listdir(".")
w = openpyxl.Workbook()
table = w.create_sheet(title="password")#在将写的文件创建sheet
 
 
afteritems = os.listdir(".")
for i in range(len(afteritems)):
    if os.path.isfile(afteritems[i]) and afteritems[i].startswith('fileName.txt'):
        file_object = open(afteritems[i], "rU")
        for (num, line) in enumerate(file_object):
            print "第 %d 行:" % (num) +line
            #读取每一行的内容，根据个人内容及需求将内容写入excelName表格中
            ssid_pass = line.split('        ')
            ssid = ssid_pass[0].split(":")
            password = ssid_pass[1].split(":")
            table.cell(num + 1, 1, ssid[1]) #num + 1 表示第几行，1 表示第几列
            table.cell(num + 1, 2, password[1])
        break
w.save(time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time())) + " excelName.xls")#根据时间来保存文件名
